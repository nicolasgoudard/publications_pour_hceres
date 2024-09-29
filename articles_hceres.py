'''
Created on 06 janv. 2022

@author: nicolasgoudard
'''
import csv
import string
import requests
import unidecode
import openpyxl
import numpy as np
from tabulate import tabulate
import pandas as pd
from _ast import Or
# TRAITE LE FICHIER EXCEL EXTRAIT DE WEB OF SCIENCE :
# ON OUVRE LE FICHIER, ON Y AJOUTE  DES COLONNES DEMANDEES PAR L'HCERES 
# ET ON LE SAUVE SOUS UN AUTRE NOM

# PROCEDURE :  
# - Obtenir la liste du personnel au format csv : type, nom, prenom, equipe, orcid
# - Sortir  un fichier excel de Web OF Sciences 
# Acceder a WOS depuis l'ENT / Ressource selectroniques / Revues de A-Z   / WOS / Nouvelle Version
# Aller dans Advanced Search
# Requete 
# AD=((Inst Sci Mol Marseille OR ISM2 OR 7313 OR UMR7313 OR 6263 OR UMR6263  OR "ism 2" ) same (marseille or aix))
# AND PY=2016-2021
# cliquer sur EXPORT / Excel /  Record From= Nombre exact de resultats Record Content = Full Record 
# Ouvrir le fichier excel et enregister au format XLSX : articles_original.xlsx
# - Corriger les variables dans le script : liste_equipe et periode notamment
# - lancer le script une premiere fois : python3 articles_hceres.py
# - Dans le output recuperer les requetes DOI et AD (orcid) pour le WOS
# - Nous allons pouvoir obtenir un résultat plus précis avec le WOS
# Retourner dans WOS advanced search, faire une requete pour DOI puis pour AD, 
# Revenir dans Advaced Search et combiner les trois requetes avec OR
# Puis cliquez sur le nombre de resultats de la requete resultante 
# Telecharger un fichier articles_original.xlsx
# - Repasser le script 
# - Enfin récuperer les resultats finaux : aticles_hceres_2016_2021.xls
# Astuce : Pour connaitre les pubs hal dans le WOS non affilées à l'ism2 mais étant en fait ism2 : DOI hal (2)  not IN ISM2 (1) : publi  dans HAL pas affilé à ISM2

# fichier / colonne d'origine WOS
fichier_wos_source="articles_original.xlsx"
fichier_personnel_with_neworcids="personnel orcids.xlsx"

# liste des equipes
liste_equipes=["BiosCiences", "Chirosciences",  "CTOM", "STeRéO"]
periode_debut="2016"
periode_fin="2021"

# liste du type de personnel connu dans le fichier du personnel : en cle le type dans le fichier personnel et en valeur son role dans les publication HCERES
dict_type_personnel={"permanent" : "permanent", 
                      "doctorant" : "doctorant",
                      "postdoc" : "permanent",
                      "ATER" : "permanent",
                      "Prof. année sab" : "permanent",
                      "Prof. associé" : "permanent",
                      "prof associé" : "permanent"} 

colheader_publication_type="Publication Type"
colheader_authors="Authors"
colheader_authors_fullname="Author Full Names"
colheader_article_title="Article Title"
colheader_source_title="Source Title"
colheader_book_series_title="Book Series Title"
colheader_book_series_subtitle="Book Series Subtitle"
colheader_language="Language"
colheader_document_type="Document Type"
colheader_conference_title="Conference Title"
colheader_conference_date="Conference Date"
colheader_conference_location="Conference Location"
colheader_addresses="Addresses"
colheader_reprint_addresses="Reprint Addresses"
colheader_isbn="ISBN"
colheader_publication_annee="Publication Year"
colheader_orcids="ORCIDs"
colheader_volume="volume"
colheader_issue="issue"
colheader_start_page="Start Page"
colheader_end_page="End Page"
colheader_doi="DOI"
colheader_ut="UT (Unique WOS ID)"
colheader_openaccess= "Open Access Designations"
col_header_datasource="datasource"
col_header_source_id="source id"

# fichier cible  colonnes cible
fichier_cible="articles_hceres_{}_{}.xlsx".format(periode_debut, periode_fin)
colheader_formatted_authors_fullname="Author Full Names pour HCERES"
colheader_volume_issue="volume / issue"
colheader_pages="Pages"
colheader_interequipes="Interequipes"
colheader_doctorant_coauteur="doctorant coauteur"
colheader_corresponding_authors="premier, dernier ou correspoding auteur"
colheader_formatted_openaccess = "Open Access"

# colonnes crees pour le fichier cible 
colheader_cible_art_ouv=[colheader_authors_fullname, colheader_article_title, colheader_source_title, colheader_volume_issue,colheader_pages, colheader_publication_annee, colheader_doi, colheader_interequipes, colheader_doctorant_coauteur, colheader_corresponding_authors,colheader_formatted_openaccess,col_header_datasource,col_header_source_id] 
colheader_cible_conf=[colheader_authors_fullname, colheader_article_title, colheader_source_title, colheader_volume_issue,colheader_pages, colheader_publication_annee, colheader_doi, colheader_interequipes, colheader_conference_title, colheader_conference_date, colheader_doctorant_coauteur, colheader_corresponding_authors,colheader_formatted_openaccess,col_header_datasource,col_header_source_id] 
colheader_cible_autre=[colheader_document_type, colheader_authors_fullname, colheader_article_title, colheader_source_title, colheader_volume_issue,colheader_pages, colheader_publication_annee, colheader_doi, colheader_interequipes, colheader_book_series_title, colheader_book_series_subtitle, colheader_isbn, colheader_conference_title, colheader_conference_date, colheader_doctorant_coauteur, colheader_corresponding_authors,colheader_formatted_openaccess,col_header_datasource,col_header_source_id] 
dict_colheaders_cible_by_sheetname={"articles" : colheader_cible_art_ouv,
                              "ouvrages" : colheader_cible_art_ouv,
                              "conferences":colheader_cible_conf,
                              "autre" : colheader_cible_autre}

# association type de document WOS => feuille cible 
dict_sheetname_by_wos_doctype={ "Article":"articles", 
                  "Article; Early Access" : "articles",
                  "Review" : "articles",
                  "Editorial Material" : "articles",
                  "Review; Early Access" : "articles",
                  "Article; Book Chapter": "ouvrages",
                  "Editorial Material; Book Chapter": "ouvrages",
                  "Review; Book Chapter" : "ouvrages",
                  "Proceedings Paper" : "ouvrages",
                  "Article; Proceedings Paper" : "conferences",
                  "Meeting Abstract" : "conferences"}

# fichier excel source
try:
    wos_workbook_source = openpyxl.load_workbook(fichier_wos_source)
except:
    print("impossible d'ouvrir le fichier :", fichier_wos_source)
    exit()
wos_worksheet_source=wos_workbook_source.active
wos_num_rows_source = wos_worksheet_source.max_row 
wos_num_cols_source = wos_worksheet_source.max_column 

# fichier excel cible: on cree plusieurs feuilles
workbook_cible=openpyxl.Workbook()
workbook_cible.remove(workbook_cible.active)
workbook_cible.create_sheet("articles")
workbook_cible.create_sheet("ouvrages")
workbook_cible.create_sheet("conferences")
workbook_cible.create_sheet("autre")

# recupere un numero de colonne par rapport a son nom en ligne 1
def getnumcol_source(colheader):
    col=1
    found=False
    while (col <=wos_num_cols_source) and not found :
        cell= wos_worksheet_source.cell(column=col, row=1)
        if (cell.value != None) and (cell.value.upper()==colheader.upper()) :
            found=True
        else :
            col=col+1
    if found :
        #print ("colonne {} est la numero {}".format(colheader, col))
        return col
    else :
        print ("colonne {} non trouvee, arret du script".format(colheader))
        exit() 
 
# affiche les erreurs sous forme de tableau
def displayErrors(lignes_en_erreur) : 
    if len (lignes_en_erreur) > 0 : 
        df = pd.DataFrame(data=lignes_en_erreur)
        print(tabulate(df, headers=("ligne", "erreur"), tablefmt='psql', showindex=False))
   
# lit le fichier permanent provoenant de l 'intranet, qui a ete recuperer avec le script export_personnel.php
# lance dans le site ism2 drupal dans le module devel - execute php code
# et met le contenu du fichier dans un dictionnaire
print ("LISTE DU PERSONNEL")
with open('personnel.csv', newline='\n', encoding="utf-8-sig", errors='ignore') as csvfile:
    reader = csv.reader(csvfile, delimiter=';', quotechar='"')
    headers_personnel=next(reader)
    personnel=[dict(zip(headers_personnel,i)) for i in reader]

# affiche le fichier personnel 
df = pd.DataFrame(data=personnel)
print(tabulate(df, headers='keys', tablefmt='psql', showindex=False))

ligne = 1
lignes_en_erreur=[]
print ("lignes en erreur dans le fichier du personnel ")
for row in personnel:
    if not row["type"] : 
        lignes_en_erreur.append((ligne, " type_personnel non renseigne"))
    elif not row["type"] in dict_type_personnel.keys() :
        lignes_en_erreur.append((ligne, "type_personnel {} inconnu".format(row["type"])))
    
    if not row["nom"] : 
        lignes_en_erreur.append((ligne, "le nom n'est pas renseigne"))
    
    if not row["prenom"] : 
        lignes_en_erreur.append((ligne, "le prenom n'est pas renseigne"))
           
    if not row["equipe"] in liste_equipes : 
        lignes_en_erreur.append((ligne, "equipe {} inconnue".format(row["equipe"]) ))
    ligne = ligne + 1

displayErrors(lignes_en_erreur)
    
# INITIALISATION DES NUMERO DE COLONNES PROVENANT DU FICHER WOS
# verifie que les colonens existent et initialise leurs numero 
print ("VERIFICATION DES COLONNES EXISTANTES : ")
col_publication_type=getnumcol_source(colheader_publication_type)
col_authors=getnumcol_source(colheader_authors)
col_authors_fullname=getnumcol_source(colheader_authors_fullname)
col_article_title=getnumcol_source(colheader_article_title)
col_source_title=getnumcol_source(colheader_source_title)
col_book_series_title=getnumcol_source(colheader_book_series_title)
col_book_series_subtitle=getnumcol_source(colheader_book_series_subtitle)
col_language=getnumcol_source(colheader_language)
col_document_type=getnumcol_source(colheader_document_type)
col_conference_title=getnumcol_source(colheader_conference_title)
col_conference_date=getnumcol_source(colheader_conference_date)
col_conference_location=getnumcol_source(colheader_conference_location)
col_addresses=getnumcol_source(colheader_addresses)
col_reprint_addresses=getnumcol_source(colheader_reprint_addresses)
col_isbn=getnumcol_source(colheader_isbn)
col_orcids=getnumcol_source(colheader_orcids)
col_publication_annee=getnumcol_source(colheader_publication_annee)
col_volume=getnumcol_source(colheader_volume)
col_issue=getnumcol_source(colheader_issue)
col_start_page=getnumcol_source(colheader_start_page)
col_end_page=getnumcol_source(colheader_end_page)
col_doi=getnumcol_source(colheader_doi)
col_ut=getnumcol_source(colheader_ut)
col_openaccess=getnumcol_source(colheader_openaccess)

# ajoutecles en-tetes des nouvelles colonnes dans chaque feuille
for sheetname in workbook_cible.sheetnames : 
    colheaders_cible=dict_colheaders_cible_by_sheetname.get(sheetname)
    for i in range(len(colheaders_cible)) : 
        workbook_cible[sheetname].cell(column=i+1, row=1).value = colheaders_cible[i] 

# PARCOURS DU FICHIER WOS
# parcours les lignes fichier excel : une ligne = une publication 
print ("TRAITEMENT DU FICHIER EXCEL WOS", fichier_wos_source)
print( 'nombre de lignes =', wos_num_rows_source, 'nombre de colonnes =', wos_num_cols_source )
lignes_en_erreur=[]
# liste de tous les orcids decouverts, et ceux des nouveaux qui ne sont pas references dans le fichier du personnel
dict_labo_tous_orcids_decouverts={}
dict_labo_nouveaux_orcids_decouverts={}
# liste des doi connus dans le wos, et des titres sans ponctuations ni espaces, qui serviront a dedoublonner wos et hal
dict_wos_doi={}
dict_wos_hash={}
for num_row in range(2, wos_num_rows_source + 1):
    #print("traitement de la ligne", num_row)
    # sur chacune des lignes on recupere une liste d'auteurs nom prenom dans la cellule et on les parcourt
    # on met en forme la liste des auteurs  nom1, prenom1; nom2 prenom2,... en NOM1 prenom1, NOM2, prenom2...
    # on repere les collaborations interequuipe et les doctorants : pour cela verifie si un des auteur est dans la liste des personnel - equipe
    # s'il est permanent on met  sont equipe dans une  liste d'equipes dedoublonee, s'il est doctorant on indique qu'un doctorant est participant 
    list_authors_shortnames=wos_worksheet_source.cell(column=col_authors, row=num_row).value.split("; ")
    list_authors_fullnames=wos_worksheet_source.cell(column=col_authors_fullname, row=num_row).value.split("; ")
    list_corresponding_authors=wos_worksheet_source.cell(column=col_reprint_addresses, row=num_row).value.split(" (corresponding author)")[0].split("; ")
    # decoupe la cellule orcid qui est de la forme, pour en extraire les champs : nom, prenom/orcid ...
    # et les mettre dans un dictionnarie indexé ainsi : (NOM, PRENOM) => orcid
    dict_authors_orcids={}
    if wos_worksheet_source.cell(column=col_orcids, row=num_row).value :
        for token in wos_worksheet_source.cell(column=col_orcids, row=num_row).value.split("; "):
            try :
                nomprenom, orcid = token.split("/") 
                nom,prenom= nomprenom.upper().split(", ")
                dict_authors_orcids[(nom,prenom)]=orcid
            except ValueError :
                lignes_en_erreur.append((num_row, "Champ orcids malformé, impossible d'extraire les informations du token {},".format(token)))
                             
    interequipe=[];
    formatted_authors_fullname=""
    doctorant_coauteur="N"
    corresponding_author="N"
    comma=""
    for indice_author_fullname in range(len(list_authors_fullnames)) : 
        author_fullname=list_authors_fullnames[indice_author_fullname]
        author_lastname="" 
        author_firstname=""
        if ", " in author_fullname :
            try :
                author_lastname, author_firstname= author_fullname.split(", ")
            except ValueError :
                lignes_en_erreur.append((num_row, "Champ author_fullname malformé, impossible d'extraire les informations du token {},".format(author_fullname)))
                values=author_fullname.split(", ")
                author_lastname=values[0]
                if len(values) > 1 :
                    author_firstname=" ".join(values[1:])
        else :
            author_lastname = author_fullname
        formatted_authors_fullname= formatted_authors_fullname + comma + author_lastname.upper() + " " + author_firstname
        author_shortname = list_authors_shortnames[indice_author_fullname]
        # pour chacun des auteurs de la publi on parcourt le fichier du personnel : si c'est un personnel de lism2, on recupere son equipe dans une liste dedoublonner et 
        # on verifie si l'auteur est doctorant coauteur / corresponding auteur ou dernier ou premier auteur
        # on extrait son orcid de la colonnes orcids pour le conserver sur un intranet perso et construire une requete plus affinée 
        comma=", "
        for row in personnel:
            personnel_nom_upper=unidecode.unidecode(row["nom"].upper());
            personnel_prenom_upper=unidecode.unidecode(row["prenom"].upper())
            if ( (personnel_nom_upper == author_lastname.upper())  and (personnel_prenom_upper == author_firstname.upper() ) ) or  ((personnel_nom_upper == author_firstname.upper()) and (personnel_prenom_upper == author_lastname.upper()) ) :
                # assertion, l'auteur fait partie de l'ism2
                type_personnel=dict_type_personnel.get(row["type"]);
                if type_personnel =="permanent" :
                    if not row["equipe"] in interequipe :
                        interequipe.append(row["equipe"] )
                elif type_personnel=="doctorant"  :
                    doctorant_coauteur="O"
                if (indice_author_fullname==0) or (indice_author_fullname==(len(list_authors_fullnames) -1)) or (author_shortname in list_corresponding_authors) :
                    corresponding_author="O"
                # on tente de  determiner le orcid grace au wos
                # si on devcouvre un orcid different de celui du fichier du personnel on le sauvegarde
                orcid=dict_authors_orcids.get((personnel_nom_upper, personnel_prenom_upper))
                if orcid and not dict_labo_tous_orcids_decouverts.get((personnel_nom_upper, personnel_prenom_upper)) :
                    dict_labo_tous_orcids_decouverts[(personnel_nom_upper, personnel_prenom_upper)]=orcid
                    if orcid != row["orcid"] :
                        dict_labo_nouveaux_orcids_decouverts[(personnel_nom_upper, personnel_prenom_upper)]=orcid
    #s'il ya plus d'une equipe dans la liste interequipe dedoublonee, on enregistre une colaboration interequipe
    formatted_interequipes=""
    comma=""
    for equipe in interequipe :
        try:
            formatted_interequipes = formatted_interequipes + comma + str(liste_equipes.index(equipe) + 1)
            comma=", "   
        except ValueError :
            lignes_en_erreur.append((num_row, "L'équipe {} existe pas".format(equipe)))

    # on prepare la colonne issue, pages, open access
    volume=wos_worksheet_source.cell(column=col_volume, row=num_row).value
    issue=wos_worksheet_source.cell(column=col_issue, row=num_row).value
    volume_issue="{}({})".format(volume, issue).replace("()", "")
    start_page=wos_worksheet_source.cell(column=col_start_page, row=num_row).value
    end_page=wos_worksheet_source.cell(column=col_end_page, row=num_row).value
    pages=" - ".join(map(str,[start_page, end_page]))

    # open access à O, s'il y a le mot green dans la colonne open access designation de wos
    if ( "GREEN" in wos_worksheet_source.cell(column=col_openaccess, row=num_row).value.upper()) :
        openaccess="O"
    else :
        openaccess="N"
    
    #autres colonnes necessaires
    # document type
    document_type= wos_worksheet_source.cell(column=col_document_type, row=num_row).value
    # titre produit
    titre_produit= wos_worksheet_source.cell(column=col_article_title, row=num_row).value
    # nom revue ou publication
    nom_revue_ou_pub=wos_worksheet_source.cell(column=col_source_title, row=num_row).value
    # nom revue ou publicaiton
    annee=wos_worksheet_source.cell(column=col_publication_annee, row=num_row).value
    # reference
    doi=wos_worksheet_source.cell(column=col_doi, row=num_row).value
    # titre conference
    conference_title=wos_worksheet_source.cell(column=col_conference_title, row=num_row).value
    # date conference
    conference_date=wos_worksheet_source.cell(column=col_conference_date, row=num_row).value
    #  book series title
    book_series_title=wos_worksheet_source.cell(column=col_book_series_title, row=num_row).value
    # book seriesubtilte
    book_series_subtitle=wos_worksheet_source.cell(column=col_book_series_subtitle, row=num_row).value
    # isbn
    isbn=wos_worksheet_source.cell(column=col_isbn, row=num_row).value
    # wos id 
    ut=wos_worksheet_source.cell(column=col_ut, row=num_row).value
    
    titre_hash = unidecode.unidecode(titre_produit.translate(str.maketrans('','', string.punctuation + string.whitespace)).upper())
   
    # apres l'analyse de colonnes du fichier source et la creation de valeurs formatees pour l'hceres
    # on va les copier dans le fichier cible et dans la feuille adequate : 
    # on determine la feuille excel cible (articles, ouvrages,... autre) en fonction du type de doncument en vue d'y copier la donnees 
    # si le type de document n'est pas connu, on va les mettre dans la feuille autre
    sheetname=dict_sheetname_by_wos_doctype.get(document_type)
    if not sheetname :
        sheetname="autre"
    # on affiche les colonnes dans le fichier cible. Les colonnes a afficher ne sont pas les memes selon le type de document
    colvalues_cible_art_ouv=[formatted_authors_fullname,titre_produit,nom_revue_ou_pub,volume_issue,pages,annee,doi,formatted_interequipes,doctorant_coauteur,corresponding_author,openaccess, "WOS", ut ]
    colvalues_cible_conf=[formatted_authors_fullname,titre_produit,nom_revue_ou_pub,volume_issue,pages,annee,doi,formatted_interequipes,conference_title,conference_date,doctorant_coauteur,corresponding_author,openaccess,"WOS",  ut]
    colvalues_cible_autre=[document_type, formatted_authors_fullname,titre_produit,nom_revue_ou_pub,volume_issue,pages,annee,doi,formatted_interequipes, book_series_title, book_series_subtitle,isbn, conference_title,conference_date,doctorant_coauteur,corresponding_author,openaccess,"WOS", ut]
    dict_colvalues_cible_by_sheetname={"articles" : colvalues_cible_art_ouv,
                              "ouvrages" : colvalues_cible_art_ouv,
                              "conferences":colvalues_cible_conf,
                              "autre" : colvalues_cible_autre} 
    num_rows_cible=workbook_cible[sheetname].max_row
    colvalues_cibles=dict_colvalues_cible_by_sheetname.get(sheetname)
    for i in range(len(colvalues_cibles)) :
        workbook_cible[sheetname].cell(column= i + 1, row=num_rows_cible+1).value=colvalues_cibles[i]
    
    if doi != "" :
        dict_wos_doi[(sheetname, doi)] = ut
    if titre_hash != "" :
        dict_wos_hash[(sheetname, titre_hash)] = ut
    
    
# affiche les erreurs
displayErrors(lignes_en_erreur)

if len (dict_labo_tous_orcids_decouverts) > 0 : 
    # On complete le fichier excel du personnel avec leur orcid decouverts 
    # Pour cela on cree une liste a partir du dictionnaire personnel, a laquelle on rajoute les orcid pour chacun des agents
    # et ensuite on cree un nouveau fichier excel a partir de cette liste 
    # on construit aussi la requete WOS sur les orcids mais uniquement pour les permanents
    wbOrcids = openpyxl.Workbook()
    ws = wbOrcids.active
    headers_personnel_neworcids = headers_personnel + ["orcid découverts"]
    listToSave = [headers_personnel_neworcids]
    req_orcid_wos="AI=("
    separator=""
    for row in personnel:
        orcid=dict_labo_tous_orcids_decouverts.get((unidecode.unidecode(row["nom"].upper()), unidecode.unidecode(row["prenom"].upper())))
        if orcid :
            xlsLine=list(row.values())
            xlsLine.append(orcid)
            # pour la requete ORCID on veut reellement les permanents pour ne pas risquer de prendre les publications 
            # des non permanents qui n'etaient pas a l'ism2 dans la periode  
            if row["type"] == "permanent":
                req_orcid_wos=req_orcid_wos+ separator + orcid
                separator=" OR "
        else :
            xlsLine=list(row.values())
        listToSave.append(xlsLine) 
    req_orcid_wos=req_orcid_wos+") AND PY={}-{}".format(periode_debut,periode_fin)
    print ("requete sur ORCID pour WOS")
    print (req_orcid_wos)
    for i in range(len(listToSave)) :
        for j in range(len(listToSave[i])) :
            ws.cell(column=j+1, row=i+1).value = listToSave[i][j]
    ws.cell(column=1, row=ws.max_row+2).value = "requete à effectuer sur les ORCID pour WOS :"
    ws.cell(column=1, row=ws.max_row+1).value =req_orcid_wos
    
# orcid decouverts : si on a decouvert des orcids
if len (dict_labo_nouveaux_orcids_decouverts) > 0 :   
    wbOrcids.save(fichier_personnel_with_neworcids)    
    print ("veuillez consulter le fichier des orcids decouverts nommé : ", fichier_personnel_with_neworcids)
    # affichage des orcid a l'ecran on transforme le dict orcids en list pour compatibilite avec DataFrame
    list_labo_orcids_decouverts =[ (nomprenom[0] + " " + nomprenom[1] , orcid) for nomprenom, orcid in dict_labo_nouveaux_orcids_decouverts.items() ]
    print ("LISTE DES ORCIDs DECOUVERTS ET NON REFERENCES DANS L'INTRANET :")
    df = pd.DataFrame(data=list_labo_orcids_decouverts)
    print(tabulate(df, headers=("nom prénom", "orcid"), tablefmt='psql', showindex=False))
    
# ICI COMMENCE LE SCRIPT HAL  
lignes_en_erreur=[]
# association type de document WOS => feuille cible 
dict_sheetname_by_hal_doctype={ "ART":"articles", 
                  "OUV": "ouvrages",
                  "COUV" : "ouvrages",
                  "COMM" : "conferences"}

hal_common_fields="halId_s,docType_s,authLastName_s,authFirstName_s,label_s,title_s,journalTitle_s,producedDate_tdate,producedDateY_i,volume_s,issue_s,page_s,doiId_s,label_xml,file_main,halId_s,files_s,linkExtUrl_s"
hal_conf_fields="conferenceStartDate_s,conferenceTitle_s,conferenceEndDate_s,conferenceOrganizer_s,city_s,country_s"
hal_book_fields="bookTitle_s,subTitle_s,isbn_s"
hal_fields="{},{}".format(hal_common_fields, hal_conf_fields)
hal_conditions="producedDateY_i:[{} TO {}]".format(periode_debut,periode_fin)
hal_filters="structId_i:186403"
#requete="https://api.archives-ouvertes.fr/search/UNIV-AMU/?q=*:*&wt=json&rows=10000&sort=producedDate_tdate%20desc&fl={}&fq={}"
hal_requete="http://api.archives-ouvertes.fr/search/?q={}&wt=json&rows=10000&fl={}&fq={}".format(hal_conditions,hal_fields,hal_filters)
print ("requete HAL :", hal_requete)
r=requests.get(hal_requete)
response_dec = r.json()
hal_number_of_docs = response_dec["response"]["numFound"]
hal_docs = response_dec["response"]["docs"]

print ("nombre de documents trouves :", hal_number_of_docs, "- nombre de documents recuperes :", len(hal_docs))

lst_doctype=[]
hal_number_of_docs=[]
req_doi_pour_wos=""
for num_row, hal_doc in enumerate(hal_docs):
    # recuperations des champs et initialisation des variables
    
    # halId
    if "halId_s" in hal_doc:
        halId=hal_doc["halId_s"]
    else :
        halId=""
        
    # type document
    if "docType_s" in hal_doc:
        doctype=hal_doc["docType_s"]
    else :
        doctype=""
       
    # titre article
    if "title_s" in hal_doc :
        article_title=hal_doc["title_s"][0]
    else :
        article_title=""
       
    # journal
    if "journalTitle_s" in hal_doc:
        journal=hal_doc["journalTitle_s"]
    else :
        journal=""
 
    # volume
    if "volume_s" in hal_doc :
        volume=hal_doc["volume_s"]
    else :
        volume=""
    
    # issue
    if "issue_s" in hal_doc :
        issue=hal_doc["issue_s"][0]
    else :
        issue=""

    # annee
    if "producedDateY_i" in hal_doc :
        annee=hal_doc["producedDateY_i"]
    else :
        annee=""
    
    # pages
    if "page_s" in hal_doc  :
        pages=hal_doc["page_s"]
    else :
        pages =""
        
    # doi
    if "doiId_s" in hal_doc  :
        doi=hal_doc["doiId_s"]
    else :
        doi =""
        
    # isbn
    if "isbn_s" in hal_doc  :
        isbn=hal_doc["isbn_s"]
    else :
        isbn =""
    
    # titre de l'ouvrage
    if "bookTitle_s" in hal_doc  :
        book_title=hal_doc["bookTitle_s"]
    else :
        book_title =""
    
    # sous-titre
    if "subTitle_s" in hal_doc  :
        subtitle=hal_doc["subTitle_s"]
    else :
        subtitle =""
        
    # date conference 
    if "conferenceStartDate_s" in hal_doc  :
        conference_date=hal_doc["conferenceStartDate_s"]
    else :
        conference_date =""
    
    # nom de la conference
    if "conferenceTitle_s" in hal_doc  :
        conference_title=hal_doc["conferenceTitle_s"]
    else :
        conference_title =""   
    
    titre_hash = unidecode.unidecode(article_title.translate(str.maketrans('','', string.punctuation + string.whitespace)).upper())
  
    if doi != "" :
        if req_doi_pour_wos != "" : 
            req_doi_pour_wos = req_doi_pour_wos + " OR " + doi
        else :
            req_doi_pour_wos =  doi


    # on parcourt le tableau hal de noms et celui de prenoms pour en faire une chaine formatee pour hceres
    # NOM1 Prenom1, NOM2, Prenom2, etc... on determine aussi les equipes qui ont colaboree, en fonction des auteurs
    # et aussi si un des doctorants et s'il y a un premier ou dernier auteur (corresponding)
    interequipe=[];
    formatted_authors_fullname=""
    doctorant_coauteur="N"
    corresponding_author="?"
    comma=""
    for i in range(len(hal_doc["authLastName_s"])) :
        formatted_authors_fullname = formatted_authors_fullname + comma + hal_doc["authLastName_s"][i].upper() + " " + hal_doc["authFirstName_s"][i]
        comma=", "
        for row in personnel:
            personnel_nom_upper=unidecode.unidecode(row["nom"].upper());
            personnel_prenom_upper=unidecode.unidecode(row["prenom"].upper())
            if ((personnel_nom_upper == unidecode.unidecode(hal_doc["authLastName_s"][i].upper())) and (personnel_prenom_upper == unidecode.unidecode(hal_doc["authFirstName_s"][i].upper())))  or  ((personnel_nom_upper == unidecode.unidecode(hal_doc["authFirstName_s"][i].upper())) and (personnel_prenom_upper == unidecode.unidecode(hal_doc["authLastName_s"][i].upper()))) :
                # assertion, l'auteur fait partie de l'ism2
                type_personnel=dict_type_personnel.get(row["type"]);
                if type_personnel=="permanent" :
                    if not row["equipe"] in interequipe :
                        interequipe.append(row["equipe"] )
                elif type_personnel=="doctorant"  :
                    doctorant_coauteur="O"
                if (i==0) or (i==(len(hal_doc["authLastName_s"]) -1)) :
                    corresponding_author="O"
    
    #s'il y a plus d'une equipe dans la liste interequipe dedoublonee, on enregistre une colaboration interequipe
    formatted_interequipes=""
    comma=""
    for equipe in interequipe :
        try:
            formatted_interequipes = formatted_interequipes + comma + str(liste_equipes.index(equipe) + 1)
            comma=", "   
        except ValueError :
            lignes_en_erreur.append((num_row, "L'équipe {} existe pas".format(equipe)))
            
    
    volume_issue="{}({})".format(volume, issue).replace("()", "")  
    
    # open access  velur=O si le champ link ou file est renseignee
    if ("linkExtUrl_s" in hal_doc) or ("files_s" in hal_doc) :
        openaccess="O"
    else :
        openaccess="N" 
    
    sheetname=dict_sheetname_by_hal_doctype.get(doctype)
    if not sheetname :
        sheetname="autre"
    
    # passe a l'enregistement suivant si la publi courante hal a deja ete extraite du WOS
    if sheetname != "autre" :
        ut=dict_wos_doi.get((sheetname, doi))
        if (doi != "") and  ut :
            lignes_en_erreur.append((num_row, "{}. HAL doi {} deja dans WOS (id={}), ignore l'enregistement HAL".format(halId, doi, ut)))
            continue
        ut=dict_wos_hash.get((sheetname, doi))
        if (titre_produit != "") and ut :
            lignes_en_erreur.append((num_row, "{}. HAL titre {} deja dans WOS  (id={}), ignore l'enregistement HAL".format(halId, titre_produit, ut)))
            continue
    
    # on ajoute les colonnes demandes par l'HCERES 
    colvalues_cible_art_ouv=[formatted_authors_fullname,article_title,journal,volume_issue,pages,annee,doi,formatted_interequipes,doctorant_coauteur,corresponding_author,openaccess,"HAL " , halId]
    colvalues_cible_conf=[formatted_authors_fullname,article_title,journal,volume_issue,pages,annee,doi,formatted_interequipes,conference_title,conference_date,doctorant_coauteur,corresponding_author,openaccess,"HAL",halId ]
    colvalues_cible_autre=[doctype, formatted_authors_fullname,article_title,journal,volume_issue,pages,annee,doi,formatted_interequipes, book_title, subtitle, isbn, conference_title,conference_date,doctorant_coauteur,corresponding_author,openaccess,"HAL", halId]
    dict_colvalues_cible_by_sheetname={"articles" : colvalues_cible_art_ouv,
                                  "ouvrages" : colvalues_cible_art_ouv,
                                  "conferences":colvalues_cible_conf,
                                  "autre" : colvalues_cible_autre} 
    num_rows_cible=workbook_cible[sheetname].max_row
    colvalues_cibles=dict_colvalues_cible_by_sheetname.get(sheetname)
    for i in range(len(colvalues_cibles)) :
        workbook_cible[sheetname].cell(column= i + 1, row=num_rows_cible+1).value=colvalues_cibles[i]
  
# affiche les logs hal
displayErrors(lignes_en_erreur)

# requete doi pour OS
print ("requete DOI pour WOS si besoin : DO=({})" .format(req_doi_pour_wos))

# FEUILLE EXCEL DE LA LISTE DES EQUIPES
#inserer une feuille avec les equpes et leur numero
workbook_cible.create_sheet("equipes")
workbook_cible["equipes"]["A1"]= "Numero de l'equipe"
workbook_cible["equipes"]["B1"]= "Nom de l'equipe"
for i in range(len(liste_equipes)) :
    workbook_cible["equipes"].cell(column=1, row=i+2).value = i + 1
    workbook_cible["equipes"].cell(column=2, row=i+2).value = liste_equipes[i]

# sauve le fichier
workbook_cible.save(fichier_cible)   

print ("FIN : veuillez consulter le fichier final pour l'HCERES", fichier_cible)
